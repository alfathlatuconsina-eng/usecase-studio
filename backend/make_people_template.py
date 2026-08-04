#!/usr/bin/env python3
"""
Generate people_data_template.xlsx  — a ready-to-fill import template.
Run once:  venv/bin/python make_people_template.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WB_PATH = "/tmp/people_data_template.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SAMPLE_FILL = PatternFill("solid", fgColor="EAF0FB")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def style_data(cell):
    cell.fill = SAMPLE_FILL
    cell.alignment = Alignment(vertical="center", wrap_text=False)
    cell.border = BORDER


def make_trainings_sheet(wb):
    ws = wb.create_sheet("Trainings")
    ws.row_dimensions[1].height = 30

    headers = [
        ("Name", 36, "Full training / program name (required)"),
        ("Category", 18, "Compliance | Technical | Leadership | Regulatory | Soft Skills"),
        ("Method", 14, "Online | Offline | Blended"),
        ("Organizer", 24, "e.g. Internal HR / vendor name"),
        ("Date Start", 13, "YYYY-MM-DD"),
        ("Date End", 13, "YYYY-MM-DD"),
        ("Target Pax", 12, "Number of planned participants"),
        ("Actual Pax", 12, "Number who actually attended"),
        ("Status", 13, "Planned | In Progress | Completed"),
        ("Budget", 14, "Total budget (IDR, number only)"),
        ("Realization", 14, "Actual spend (IDR, number only)"),
        ("Notes", 30, "Optional free text"),
    ]

    for col, (title, width, comment) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = width

    samples = [
        ["Servant Leadership Workshop", "Leadership", "Offline", "MNC Internal HR",
         "2026-07-14", "2026-07-15", 20, None, "Planned", 15000000, None, ""],
        ["Anti-Fraud Awareness", "Compliance", "Online", "OJK E-Learning",
         "2026-08-01", "2026-08-31", 150, None, "Planned", 5000000, None, "Mandatory for all staff"],
    ]
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_data(cell)

    ws.freeze_panes = "A2"
    return ws


def make_evaluations_sheet(wb):
    ws = wb.create_sheet("Evaluations")
    ws.row_dimensions[1].height = 30

    headers = [
        ("Training Name", 40, "Must exactly match a name in the Trainings sheet or in the DB"),
        ("Reaction Score", 16, "Kirkpatrick L1 — scale 1.0 to 5.0"),
        ("Learning Score", 16, "Kirkpatrick L2 — scale 0 to 100"),
        ("Respondents", 14, "Number of survey respondents"),
        ("Notes", 30, "Optional free text"),
    ]

    for col, (title, width, comment) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = width

    samples = [
        ["Servant Leadership Workshop", 4.3, 80.0, 18, ""],
        ["Anti-Fraud Awareness", 3.9, 74.5, 120, "Post-training quiz"],
    ]
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_data(cell)

    ws.freeze_panes = "A2"
    return ws


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default Sheet
    make_trainings_sheet(wb)
    make_evaluations_sheet(wb)
    wb.save(WB_PATH)
    print(f"Template saved to: {WB_PATH}")


if __name__ == "__main__":
    main()
