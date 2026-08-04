#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Import People Development data from DB Training.xlsx into 4 tables:
  1__DB_Training        → pd_training
  2__Evaluate_Event     → pd_evaluate_event
  3__Evaluate_Facilitator → pd_evaluate_facilitator
  4__Ikatan_Dinas       → pd_ikatan_dinas

Usage:
    cd /opt/pmo/backend
    source venv/bin/activate
    python import_pd_excel.py /tmp/people_data.xlsx

Safe to re-run: existing rows are never modified or deleted (append-only).
Duplicate detection uses program_code + participant_nip for Training,
program_code for Evaluate sheets, and no + participant_nip for Ikatan Dinas.
"""

import sys
import datetime as dt
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not found — run: pip install openpyxl")

from app import Session, PdTraining, PdEvaluateEvent, PdEvaluateFacilitator, PdIkatanDinas
from normalize import normalize_directorate, normalize_classification
from sqlalchemy import select, tuple_, event
from sqlalchemy import String as SAString


# ---------------------------------------------------------------------------
# Safety net for messy source data: some rows have values landing in the
# wrong columns (e.g. a venue address in a date field). Rather than let one
# oversized string abort the whole import, truncate any string to its
# column's declared max length right before insert.
# ---------------------------------------------------------------------------

def _truncate_before_insert(mapper, connection, target):
    for col in mapper.columns:
        if isinstance(col.type, SAString) and col.type.length:
            val = getattr(target, col.key, None)
            if isinstance(val, str) and len(val) > col.type.length:
                setattr(target, col.key, val[:col.type.length])


for _model in (PdTraining, PdEvaluateEvent, PdEvaluateFacilitator, PdIkatanDinas):
    event.listen(_model, "before_insert", _truncate_before_insert)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _str(v, default=""):
    if v is None:
        return default
    s = str(v).strip()
    return default if s in ("#REF!", "#N/A", "#VALUE!", "#DIV/0!") else s


def _int(v):
    if v is None or _str(v) == "":
        return None
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return None


def _float(v):
    if v is None or _str(v) == "":
        return None
    try:
        return float(str(v))
    except (ValueError, TypeError):
        return None


def _date(v):
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s


def _time(v):
    if v is None:
        return ""
    if isinstance(v, dt.time):
        return v.strftime("%H:%M")
    if isinstance(v, dt.datetime):
        return v.strftime("%H:%M")
    return _str(v)


def rows_from(ws, min_row):
    return ws.iter_rows(min_row=min_row, values_only=True)


# ---------------------------------------------------------------------------
# 1__DB_Training  (header row 1, data from row 2)
# ---------------------------------------------------------------------------

def import_training(ws, session):
    # col indices (0-based) from row 1
    H = {str(c.value).strip().lower(): c.column - 1
         for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value}

    def g(row, *keys):
        for k in keys:
            i = H.get(k.lower())
            if i is not None and i < len(row):
                return row[i]
        return None

    # Dedup on (program_code, participant_nip, start_date, end_date):
    # program_code is a recycled monthly-slot code (e.g. "Apr001"), reused every
    # year, so (code, nip) alone wrongly merges a participant's distinct sessions
    # across years. Including the dates keeps real sessions while still dropping
    # exact duplicates.
    existing = set(session.execute(
        select(PdTraining.program_code, PdTraining.participant_nip,
               PdTraining.start_date, PdTraining.end_date)
    ).all())

    BATCH = 500
    added = skipped = 0
    for row in rows_from(ws, 2):
        if all(v is None for v in row):
            continue
        code = _str(g(row, "program code"))
        nip  = _str(g(row, "participant nip"))
        start = _date(g(row, "start date"))
        end   = _date(g(row, "end date"))
        key = (code, nip, start, end)
        if key in existing:
            skipped += 1
            continue

        session.add(PdTraining(
            program_code          = code,
            program_title         = _str(g(row, "program title")),
            program_sub_title     = _str(g(row, "program sub title")),
            training_type         = _str(g(row, "training type for dashboard")),
            flag_calculation      = _str(g(row, "flag for calculation")),
            group_training_type   = _str(g(row, "group training type")),
            bank_training_type    = _str(g(row, "bank training type")),
            training_classification = normalize_classification(_str(g(row, "training classification"))),
            training_method       = _str(g(row, "training method")),
            source_implement      = _str(g(row, "source implement")),
            institution           = _str(g(row, "institution")),
            venue                 = _str(g(row, "venue")),
            start_date            = start,
            end_date              = end,
            month                 = _str(g(row, "month")),
            year                  = _int(g(row, "year")),
            days                  = _float(g(row, "days")),
            hours_per_day         = _float(g(row, "training \n(hours in a day)", "training (hours in a day)", "hours in a day")),
            total_training_hours  = _float(g(row, "total training hours")),
            man_days              = _float(g(row, "man-days")),
            participant_nip       = nip,
            participant_name      = _str(g(row, "participant name")),
            grade                 = _int(g(row, "grade")),
            grade_description     = _str(g(row, "grade description")),
            position              = _str(g(row, "position")),
            kelompok_posisi       = _str(g(row, "kelompok posisi")),
            gender                = _str(g(row, "gender")),
            layer                 = _str(g(row, "layer")),
            branch                = _str(g(row, "branch")),
            city                  = _str(g(row, "city")),
            group_name            = _str(g(row, "group")),
            directorate           = normalize_directorate(_str(g(row, "directorate"))),
        ))
        existing.add(key)
        added += 1
        if added % BATCH == 0:
            session.commit()
            session.expunge_all()
            print(f"    ... committed {added} rows", flush=True)

    return added, skipped


# ---------------------------------------------------------------------------
# 2__Evaluate_Event  (headers span rows 3–4, data from row 5)
# ---------------------------------------------------------------------------

def import_evaluate_event(ws, session):
    # Row 3: top-level labels; Row 4: sub-labels
    r3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    r4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    # Build column map by combining rows 3 and 4
    col_map = {}
    last_top = ""
    for i, (top, sub) in enumerate(zip(r3, r4)):
        top = _str(top) or last_top
        if top:
            last_top = top
        sub = _str(sub)
        label = (top + " " + sub).strip().lower() if sub else top.lower()
        col_map[label] = i

    def g(row, *keys):
        for k in keys:
            i = col_map.get(k.lower())
            if i is not None and i < len(row):
                return row[i]
        return None

    # Dedup on (program_code, start_date, end_date): program_code is a recycled
    # monthly-slot code, so keying on it alone merges the same slot's events
    # across years into one.
    existing = set(session.execute(
        select(PdEvaluateEvent.program_code,
               PdEvaluateEvent.start_date, PdEvaluateEvent.end_date)
    ).all())

    BATCH = 500
    added = skipped = 0
    for row in rows_from(ws, 5):
        if all(v is None for v in row):
            continue
        code = _str(row[0]) if row[0] else ""
        if not code:
            skipped += 1
            continue
        start = _date(row[3]) if len(row) > 3 else ""
        end   = _date(row[4]) if len(row) > 4 else ""
        key = (code, start, end)
        if key in existing:
            skipped += 1
            continue

        session.add(PdEvaluateEvent(
            program_code        = code,
            program_title       = _str(row[1]) if len(row) > 1 else "",
            program_sub_title   = _str(row[2]) if len(row) > 2 else "",
            start_date          = start,
            end_date            = end,
            start_time          = _time(row[5]) if len(row) > 5 else "",
            end_time            = _time(row[6]) if len(row) > 6 else "",
            hours_sum           = _float(row[7]) if len(row) > 7 else None,
            total_days          = _float(row[8]) if len(row) > 8 else None,
            score_latar_belakang  = _float(row[9])  if len(row) > 9  else None,
            score_struktur_materi = _float(row[10]) if len(row) > 10 else None,
            score_kemudahan       = _float(row[11]) if len(row) > 11 else None,
            score_korelasi        = _float(row[12]) if len(row) > 12 else None,
            score_case_study      = _float(row[13]) if len(row) > 13 else None,
            score_kenyamanan      = _float(row[14]) if len(row) > 14 else None,
            score_fasilitas       = _float(row[15]) if len(row) > 15 else None,
            score_konsumsi        = _float(row[16]) if len(row) > 16 else None,
        ))
        existing.add(key)
        added += 1
        if added % BATCH == 0:
            session.commit()
            session.expunge_all()
            print(f"    ... committed {added} rows", flush=True)

    return added, skipped


# ---------------------------------------------------------------------------
# 3__Evaluate_Facilitator  (headers span rows 3–4, data from row 5)
# ---------------------------------------------------------------------------

def import_evaluate_facilitator(ws, session):
    # Dedup on (program_code, facilitator_nip, start_date, end_date): program_code
    # is a recycled monthly-slot code, so keying on (code, nip) alone merges the
    # same facilitator's sessions for that slot across years into one.
    existing = set(session.execute(
        select(PdEvaluateFacilitator.program_code, PdEvaluateFacilitator.facilitator_nip,
               PdEvaluateFacilitator.start_date, PdEvaluateFacilitator.end_date)
    ).all())

    BATCH = 500
    added = skipped = 0
    for row in rows_from(ws, 5):
        if all(v is None for v in row):
            continue
        # col 1 = program_code, col 4 = facilitator_nip (0-based)
        code = _str(row[1]) if len(row) > 1 else ""
        nip  = _str(row[4]) if len(row) > 4 else ""
        if not code:
            skipped += 1
            continue
        start = _date(row[7]) if len(row) > 7 else ""
        end   = _date(row[8]) if len(row) > 8 else ""
        key = (code, nip, start, end)
        if key in existing:
            skipped += 1
            continue

        session.add(PdEvaluateFacilitator(
            program_code          = code,
            program_title         = _str(row[2])  if len(row) > 2  else "",
            program_sub_title     = _str(row[3])  if len(row) > 3  else "",
            facilitator_nip       = nip,
            facilitator_name      = _str(row[5])  if len(row) > 5  else "",
            facilitator_position  = _str(row[6])  if len(row) > 6  else "",
            start_date            = start,
            end_date              = end,
            start_time            = _time(row[9]) if len(row) > 9  else "",
            end_time              = _time(row[10])if len(row) > 10 else "",
            hours_sum             = _float(row[11])if len(row) > 11 else None,
            total_days            = _float(row[12])if len(row) > 12 else None,
            point                 = _str(row[13]) if len(row) > 13 else "",
            score_pengelolaan_waktu    = _float(row[14]) if len(row) > 14 else None,
            score_menjelaskan_materi   = _float(row[15]) if len(row) > 15 else None,
            score_pemahaman            = _float(row[16]) if len(row) > 16 else None,
            score_manajemen_interaksi  = _float(row[17]) if len(row) > 17 else None,
        ))
        existing.add(key)
        added += 1
        if added % BATCH == 0:
            session.commit()
            session.expunge_all()
            print(f"    ... committed {added} rows", flush=True)

    return added, skipped


# ---------------------------------------------------------------------------
# 4__Ikatan_Dinas  (header row 2, data from row 3)
# ---------------------------------------------------------------------------

def import_ikatan_dinas(ws, session):
    existing = set(session.execute(
        select(PdIkatanDinas.no, PdIkatanDinas.participant_nip)
    ).all())

    BATCH = 500
    added = skipped = 0
    for row in rows_from(ws, 3):
        if all(v is None for v in row):
            continue
        # col layout (0-based): 0=None,1=No,2=Classification,...
        no  = _int(row[1]) if len(row) > 1 else None
        nip = _str(row[19]) if len(row) > 19 else ""
        key = (no, nip)
        if key in existing:
            skipped += 1
            continue

        session.add(PdIkatanDinas(
            no                    = no,
            training_classification = normalize_classification(_str(row[2]) if len(row) > 2 else ""),
            allocation_group      = _str(row[3])  if len(row) > 3  else "",
            training_city         = _str(row[4])  if len(row) > 4  else "",
            institution_name      = _str(row[5])  if len(row) > 5  else "",
            material              = _str(row[6])  if len(row) > 6  else "",
            facilitator_1         = _str(row[7])  if len(row) > 7  else "",
            facilitator_2         = _str(row[8])  if len(row) > 8  else "",
            facilitator_3         = _str(row[9])  if len(row) > 9  else "",
            facilitator_4         = _str(row[10]) if len(row) > 10 else "",
            facilitator_5         = _str(row[11]) if len(row) > 11 else "",
            start_date            = _date(row[12])if len(row) > 12 else "",
            end_date              = _date(row[13])if len(row) > 13 else "",
            ikatan_dinas          = _str(row[14]) if len(row) > 14 else "",
            days                  = _float(row[15])if len(row) > 15 else None,
            hours_per_day         = _float(row[16])if len(row) > 16 else None,
            total_training_hours  = _float(row[17])if len(row) > 17 else None,
            man_days              = _float(row[18])if len(row) > 18 else None,
            participant_nip       = nip,
            participant_name      = _str(row[20]) if len(row) > 20 else "",
            job_title             = _str(row[21]) if len(row) > 21 else "",
            group_name            = _str(row[22]) if len(row) > 22 else "",
            participant_city      = _str(row[23]) if len(row) > 23 else "",
            penalty_amount        = _float(row[24])if len(row) > 24 else None,
            keterangan            = _str(row[25]) if len(row) > 25 else "",
        ))
        existing.add(key)
        added += 1
        if added % BATCH == 0:
            session.commit()
            session.expunge_all()
            print(f"    ... committed {added} rows", flush=True)

    return added, skipped


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

SHEET_MAP = {
    "1__db_training":          ("Training",             import_training),
    "2__evaluate_event":       ("Evaluate Event",       import_evaluate_event),
    "3__evaluate_facilitator": ("Evaluate Facilitator", import_evaluate_facilitator),
    "4__ikatan_dinas":         ("Ikatan Dinas",         import_ikatan_dinas),
}


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python import_pd_excel.py <path/to/file.xlsx>")

    path = Path(sys.argv[1])
    if not path.exists():
        sys.exit(f"File not found: {path}")

    print(f"Loading {path.name} ...")
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)

    total_added = 0
    with Session() as session:
        for sheet_name in wb.sheetnames:
            handler = SHEET_MAP.get(sheet_name.lower())
            if not handler:
                continue
            label, fn = handler
            print(f"\n=== {label} (sheet: {sheet_name}) ===")
            a, s = fn(wb[sheet_name], session)
            print(f"  Added: {a}  |  Skipped: {s}")
            total_added += a

        session.commit()

    wb.close()
    print(f"\nDone. Total rows inserted: {total_added}")


if __name__ == "__main__":
    main()
