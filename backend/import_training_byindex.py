#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Position-based importer for the 2024-2026 DB Training file.

That file has an extra unlabeled column at index 4 and is missing the
'Kelompok Posisi' column, so the header-name mapping in import_pd_excel.py
reads the wrong column for Year / NIP / Man-Days / dates. This script maps
by fixed column INDEX (decoded from the file) instead.

Append-only with the same dedup key as the stock importer
(program_code, participant_nip, start_date, end_date).

Usage:
    cd /opt/pmo/backend && source venv/bin/activate
    python import_training_byindex.py <file.xlsx>            # PREVIEW only
    python import_training_byindex.py <file.xlsx> --commit   # actually insert
"""

import sys
import itertools
import openpyxl

from app import Session, PdTraining
from import_pd_excel import _str, _int, _float, _date
from normalize import normalize_directorate, normalize_classification
from sqlalchemy import select

SHEET = "1__DB_Training"

# decoded column indices for THIS file (0-based). index 4 (=extra col) and
# 'kelompok_posisi' are intentionally absent.
I = {
    "program_code": 0, "program_title": 1, "program_sub_title": 2,
    "training_type": 3,
    # 4 = extra/ignored
    "flag_calculation": 5, "group_training_type": 6, "bank_training_type": 7,
    "training_classification": 8, "training_method": 9, "source_implement": 10,
    "institution": 11, "venue": 12, "start_date": 13, "end_date": 14,
    "month": 15, "year": 16, "days": 17, "hours_per_day": 18,
    "total_training_hours": 19, "man_days": 20, "participant_nip": 21,
    "participant_name": 22, "grade": 23, "grade_description": 24,
    "position": 25, "gender": 26, "layer": 27, "branch": 28, "city": 29,
    "group_name": 30, "directorate": 31,
}


def cell(row, key):
    i = I.get(key)
    return row[i] if (i is not None and i < len(row)) else None


def build(row):
    start = _date(cell(row, "start_date"))
    end = _date(cell(row, "end_date"))
    return dict(
        program_code=_str(cell(row, "program_code")),
        program_title=_str(cell(row, "program_title")),
        program_sub_title=_str(cell(row, "program_sub_title")),
        training_type=_str(cell(row, "training_type")),
        flag_calculation=_str(cell(row, "flag_calculation")),
        group_training_type=_str(cell(row, "group_training_type")),
        bank_training_type=_str(cell(row, "bank_training_type")),
        training_classification=normalize_classification(_str(cell(row, "training_classification"))),
        training_method=_str(cell(row, "training_method")),
        source_implement=_str(cell(row, "source_implement")),
        institution=_str(cell(row, "institution")),
        venue=_str(cell(row, "venue")),
        start_date=start,
        end_date=end,
        month=_str(cell(row, "month")),
        year=_int(cell(row, "year")),
        days=_float(cell(row, "days")),
        hours_per_day=_float(cell(row, "hours_per_day")),
        total_training_hours=_float(cell(row, "total_training_hours")),
        man_days=_float(cell(row, "man_days")),
        participant_nip=_str(cell(row, "participant_nip")),
        participant_name=_str(cell(row, "participant_name")),
        grade=_int(cell(row, "grade")),
        grade_description=_str(cell(row, "grade_description")),
        position=_str(cell(row, "position")),
        kelompok_posisi="",
        gender=_str(cell(row, "gender")),
        layer=_str(cell(row, "layer")),
        branch=_str(cell(row, "branch")),
        city=_str(cell(row, "city")),
        group_name=_str(cell(row, "group_name")),
        directorate=normalize_directorate(_str(cell(row, "directorate"))),
    )


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python import_training_byindex.py <file.xlsx> [--commit]")
    path = sys.argv[1]
    commit = "--commit" in sys.argv[2:]

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[SHEET]

    if not commit:
        print("=== PREVIEW (no data written) — first 5 mapped rows ===")
        for row in itertools.islice(ws.iter_rows(min_row=2, values_only=True), 5):
            if all(v is None for v in row):
                continue
            d = build(row)
            print("  year=%s month=%s nip=%s name=%s mandays=%s start=%s dir=%s type=%s method=%s" % (
                d["year"], d["month"], d["participant_nip"], d["participant_name"][:24],
                d["man_days"], d["start_date"], d["directorate"], d["training_type"],
                d["training_method"]))
        total = sum(1 for r in ws.iter_rows(min_row=2, values_only=True)
                    if not all(v is None for v in r))
        print("Total data rows in file:", total)
        print("\nLooks right? Re-run with  --commit  to import.")
        wb.close()
        return

    with Session() as s:
        existing = set(s.execute(
            select(PdTraining.program_code, PdTraining.participant_nip,
                   PdTraining.start_date, PdTraining.end_date)
        ).all())
        added = skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            d = build(row)
            key = (d["program_code"], d["participant_nip"], d["start_date"], d["end_date"])
            if key in existing:
                skipped += 1
                continue
            s.add(PdTraining(**d))
            existing.add(key)
            added += 1
            if added % 500 == 0:
                s.commit()
                s.expunge_all()
                print("    ... committed %d rows" % added, flush=True)
        s.commit()
        print("\nDone. Added: %d  |  Skipped: %d" % (added, skipped))
    wb.close()


if __name__ == "__main__":
    main()
