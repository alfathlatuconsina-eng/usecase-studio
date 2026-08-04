"""
Parser + validator untuk tiga jenis file Excel.

Alur: baca apa adanya -> normalisasi -> validasi -> hasilkan baris siap simpan.
Modul ini TIDAK menyentuh database. Ia murni mengubah file menjadi struktur
Python, supaya bisa diuji tanpa PostgreSQL.

Aturan penting yang lahir dari analisa data nyata:
  * Nomor deposito muncul dalam 3 format berbeda -> selalu dinormalisasi digit-only
  * Nominal bisa tersimpan sebagai teks ('20.000.000.000') -> dikonversi, ambigu ditolak
  * Kode cabang di file cabang berbentuk '1006 - GREEN GARDEN (KCP)' -> diambil kodenya
  * Data IT adalah export fixed-width -> seluruh teks di-trim, nama terpotong ditandai
"""
from __future__ import annotations

import datetime
import re
import unicodedata
from dataclasses import dataclass, field

import openpyxl

# --------------------------------------------------------------------------
# helper normalisasi
# --------------------------------------------------------------------------
_WS = re.compile(r"\s+")


def clean(v):
    """Rapikan teks: normalisasi unicode, buang padding spasi export fixed-width."""
    if v is None:
        return None
    s = unicodedata.normalize("NFKC", str(v)).strip()
    s = _WS.sub(" ", s)
    return s or None


def digits(v):
    """Ambil digit saja. Kunci pencocokan lintas sumber."""
    if v is None:
        return None
    s = re.sub(r"[^0-9]", "", str(v))
    return s or None


def branch_code(v):
    """'1006 - GREEN GARDEN (KCP)' -> '01006' ; 1006 -> '01006'"""
    if v is None:
        return None
    head = str(v).split(" - ")[0].strip()
    head = re.sub(r"[^0-9A-Za-z]", "", head)
    if not head:
        return None
    return head.zfill(5) if head.isdigit() else head


def as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def as_time(v):
    if isinstance(v, datetime.time):
        return v
    if isinstance(v, datetime.datetime):
        return v.time()
    return None


class AmbiguousNumber(ValueError):
    """Nominal berupa teks yang tidak bisa ditafsirkan tanpa menebak."""


def as_number(v):
    """
    Konversi nominal ke float.

    Menangani kasus nyata dari file cabang:
      1200000000            -> 1200000000.0
      '20.000.000.000'      -> 20000000000.0   (titik = pemisah ribuan)
      '3.098.612.448.63'    -> AmbiguousNumber (format rusak, jangan ditebak)
      '1.234,56'            -> 1234.56
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip().replace("Rp", "").replace(" ", "")
    if not s:
        return None

    if "," in s:                       # koma dianggap desimal (konvensi Indonesia)
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            raise AmbiguousNumber(str(v))

    if "." in s:
        parts = s.split(".")
        # pemisah ribuan yang benar: semua kelompok setelah yang pertama = 3 digit
        if all(len(p) == 3 for p in parts[1:]):
            return float("".join(parts))
        # satu titik dengan ekor bukan 3 digit -> desimal biasa
        if len(parts) == 2:
            try:
                return float(s)
            except ValueError:
                raise AmbiguousNumber(str(v))
        # lebih dari satu titik dan ekor tidak konsisten -> tidak bisa ditafsirkan
        raise AmbiguousNumber(str(v))

    try:
        return float(s)
    except ValueError:
        raise AmbiguousNumber(str(v))


# --------------------------------------------------------------------------
# struktur hasil
# --------------------------------------------------------------------------
@dataclass
class Issue:
    baris_no: int
    severity: str          # 'error' | 'warning'
    kode: str
    pesan: str
    kolom: str | None = None
    nilai: str | None = None
    branch_code: str | None = None


@dataclass
class ParseResult:
    jenis: str
    rows: list = field(default_factory=list)        # baris siap simpan (dict)
    raw: list = field(default_factory=list)         # (baris_no, payload) untuk staging
    issues: list = field(default_factory=list)

    @property
    def errors(self):
        return [i for i in self.issues if i.severity == "error"]

    @property
    def warnings(self):
        return [i for i in self.issues if i.severity == "warning"]

    @property
    def baris_ditolak(self):
        return len({i.baris_no for i in self.errors})

    @property
    def baris_valid(self):
        return len([r for r in self.rows if not r["_ditolak"]])

    def periode(self, kolom):
        vals = [r[kolom] for r in self.rows if r.get(kolom)]
        return (min(vals), max(vals)) if vals else (None, None)


# --------------------------------------------------------------------------
# 1. DATA IT - break deposito
# --------------------------------------------------------------------------
SHEET_IT = "Sheet1"
_TRUNC_NAMA = 20          # lebar tetap kolom nama di export core banking


def parse_it(path, branches: dict, settings: dict) -> ParseResult:
    """branches: {kode: {...}} ; settings: {'jam_operasional_mulai': 8, ...}"""
    res = ParseResult("it_break")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[SHEET_IT] if SHEET_IT in wb.sheetnames else wb.worksheets[0]

    jam_a = int(settings.get("jam_operasional_mulai", 8))
    jam_b = int(settings.get("jam_operasional_selesai", 16))
    mode_dup = str(settings.get("validasi_duplikat", "abaikan")).lower()

    it = ws.iter_rows(values_only=True)
    next(it, None)                                   # lewati header

    for n, r in enumerate(it, start=2):
        if r is None or r[0] is None:
            continue
        res.raw.append((n, [str(x) if x is not None else None for x in r]))
        err = []

        kode = branch_code(r[0])
        if kode not in branches:
            err.append(Issue(n, "error", "cabang_tak_dikenal",
                             f"Kode cabang {kode} tidak ada di master", "KodeCabang", str(r[0])))

        tgl_break = as_date(r[6])
        if tgl_break is None:
            err.append(Issue(n, "error", "tgl_break_kosong", "TanggalBreak kosong", "TanggalBreak"))

        try:
            nominal = as_number(r[11])
        except AmbiguousNumber as e:
            nominal = None
            err.append(Issue(n, "error", "nominal_ambigu",
                             f"NominalPencairan tidak bisa dibaca: {e}", "NominalPencairan", str(r[11])))
        if nominal is None and not err:
            err.append(Issue(n, "error", "nominal_kosong", "NominalPencairan kosong", "NominalPencairan"))

        tgl_tempat = as_date(r[4])
        tgl_jt = as_date(r[5])
        w_awal, w_akhir = as_time(r[7]), as_time(r[8])

        durasi = None
        if w_awal and w_akhir:
            durasi = (w_akhir.hour * 3600 + w_akhir.minute * 60 + w_akhir.second) - \
                     (w_awal.hour * 3600 + w_awal.minute * 60 + w_awal.second)

        nama_raw = str(r[10]) if r[10] is not None else ""
        terpotong = len(nama_raw) == _TRUNC_NAMA and not nama_raw.endswith(" ")

        flags = []
        if tgl_tempat and tgl_break and tgl_tempat > tgl_break:
            flags.append("Penempatan setelah tanggal break")
            res.issues.append(Issue(n, "warning", "tgl_tempat_masa_depan",
                                    f"Tanggal penempatan {tgl_tempat} setelah tanggal break {tgl_break}",
                                    "TanggalPenempatan", str(tgl_tempat)))
        luar_jam = bool(w_awal and (w_awal.hour < jam_a or w_awal.hour >= jam_b))
        if luar_jam:
            flags.append(f"Di luar jam {jam_a:02d}:00-{jam_b:02d}:00")
        if durasi is not None and durasi > 300:
            flags.append("Durasi proses >5 menit")
        if terpotong:
            flags.append("Nama nasabah terpotong")
            res.issues.append(Issue(n, "warning", "nama_terpotong",
                                    "Nama nasabah terpotong pada 20 karakter (export fixed-width)",
                                    "NamaPemilikRekening", nama_raw))

        for i in err + [x for x in res.issues if x.baris_no == n and x.branch_code is None]:
            i.branch_code = kode
        res.issues.extend(err)
        res.rows.append({
            "_ditolak": bool(err), "baris_no": n,
            "branch_code": kode if kode in branches else None,
            "cabang_core": clean(r[1]),
            "saldo": _safe_num(r[2]),
            "tgl_penempatan": tgl_tempat, "tgl_jatuh_tempo": tgl_jt, "tgl_break": tgl_break,
            "waktu_awal": w_awal, "waktu_akhir": w_akhir, "durasi_detik": durasi,
            "rek_pendebetan": clean(r[9]), "rek_norm": digits(r[9]),
            "nama_pemilik": clean(r[10]), "nama_terpotong": terpotong,
            "nominal": nominal, "penalti": _safe_num(r[12]) or 0,
            "mata_uang": clean(r[13]) or "IDR", "rate": _safe_num(r[14]),
            "rek_pencairan": clean(r[15]), "nama_pencairan": clean(r[16]),
            "via_perantara": "PERANTARA" in str(r[16] or "").upper(),
            "cs_id": clean(r[18]), "cs_nama": clean(r[19]),
            # kolom 20 (FLM1ID) di export ternyata salinan CustomerServiceID -> tidak dipakai
            "flm1_nama": clean(r[21]),
            "teller_id": clean(r[22]), "teller_nama": clean(r[23]),
            "flm2_id": clean(r[24]), "flm2_nama": clean(r[25]),
            "sisa_hari": (tgl_jt - tgl_break).days if tgl_jt and tgl_break else None,
            "umur_hari": (tgl_break - tgl_tempat).days if tgl_tempat and tgl_break else None,
            "break_sejati": bool(tgl_jt and tgl_break and tgl_break < tgl_jt),
            "luar_jam": luar_jam,
            "flags": flags,
        })

    wb.close()
    _tandai_duplikat(res, ("rek_norm", "nominal", "tgl_break"),
                     "Duplikat di data IT", mode=mode_dup)
    return res


def _safe_num(v):
    try:
        return as_number(v)
    except AmbiguousNumber:
        return None


# --------------------------------------------------------------------------
# 2. PENCAIRAN DEPOSITO - dari cabang
# --------------------------------------------------------------------------
SHEET_PENCAIRAN = "Pencairan Deposito - olah"
HEADER_ROW_PENCAIRAN = 3

# kata kunci penempatan kembali; sengaja konservatif
_KW_ROLL = re.compile(
    r"tempatkan kembali|penempatan kembali|depositokan kembali|roll\s*over|rollover|"
    r"\baro\b|perpanjang|on\s*call|\bdoc\b|otomatis cair", re.I)


def parse_pencairan(path, branches: dict, settings: dict) -> ParseResult:
    res = ParseResult("pencairan")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[SHEET_PENCAIRAN] if SHEET_PENCAIRAN in wb.sheetnames else wb.worksheets[0]
    tenor_roll = int(settings.get("rollover_tenor_hari", 1))
    cek_nip = bool(int(settings.get("validasi_nip", 0)))
    mode_dup = str(settings.get("validasi_duplikat", "abaikan")).lower()

    for n, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if n <= HEADER_ROW_PENCAIRAN or r is None or r[1] is None:
            continue
        res.raw.append((n, [str(x) if x is not None else None for x in r]))
        err = []

        kode = branch_code(r[2])
        if kode not in branches:
            err.append(Issue(n, "error", "cabang_tak_dikenal",
                             f"Kode cabang tidak dikenal: {r[2]}", "Kode Cabang", str(r[2])))

        tgl_input = as_date(r[3])
        if tgl_input is None:
            err.append(Issue(n, "error", "tgl_input_kosong", "Tanggal input kosong", "Tanggal (Input Data)"))

        try:
            nominal = as_number(r[9])
        except AmbiguousNumber as e:
            nominal = None
            err.append(Issue(n, "error", "nominal_ambigu",
                             f"Nominal Break tidak bisa dibaca tanpa menebak: {e}", "Nominal Break", str(r[9])))
        if nominal is None and r[9] is None:
            err.append(Issue(n, "error", "nominal_kosong", "Nominal Break kosong", "Nominal Break"))

        tgl_tempat, tgl_cair = as_date(r[6]), as_date(r[8])
        if tgl_tempat and tgl_cair and tgl_tempat > tgl_cair:
            err.append(Issue(n, "error", "tanggal_terbalik",
                             f"Tanggal penempatan {tgl_tempat} setelah tanggal pencairan {tgl_cair}",
                             "Tanggal Penempatan", str(tgl_tempat)))

        tenor = (tgl_cair - tgl_tempat).days if tgl_tempat and tgl_cair else None
        catatan = clean(r[16])
        kw = bool(catatan and _KW_ROLL.search(catatan))

        # klasifikasi arus dana
        if tenor is not None and tenor <= tenor_roll:
            arus, yakin = "Rollover / DOC", ("Tinggi" if kw else "Sedang")
        elif kw:
            arus, yakin = "Penempatan Kembali", "Sedang"
        else:
            arus, yakin = "Arus Keluar", ("Tinggi" if tenor is not None else "Rendah")

        flags, warn = [], []
        if r[4] is None:
            flags.append("Nomor deposito kosong")
            warn.append(("depnum_kosong", "Nomor deposito kosong", "Nomor Deposito", None))
        if tgl_cair and tgl_input and (tgl_cair - tgl_input).days > 30:
            flags.append("Pencairan >30 hari dari input")
            warn.append(("tgl_jauh", f"Tanggal pencairan {tgl_cair} lebih dari 30 hari dari input",
                         "Tanggal Pencairan", str(tgl_cair)))
        # Pemeriksaan NIP bisa dimatikan lewat Pengaturan (validasi_nip = 0).
        # Nilainya tetap disimpan apa adanya; yang dimatikan hanya penandaannya.
        if cek_nip:
            if r[13] is None and r[14] is None and r[15] is None:
                flags.append("Tanpa maker/checker/approver")
                warn.append(("nip_kosong", "Tidak ada maker, checker, maupun approver", "NIP", None))
            for idx, lbl in ((13, "maker"), (14, "checker"), (15, "approver")):
                if r[idx] is not None and len(str(r[idx]).strip()) != 9:
                    flags.append(f"NIP {lbl} bukan 9 digit")
                    warn.append((f"nip_format_{lbl}", f"NIP {lbl} bukan 9 digit: {r[idx]}",
                                 f"NIP ({lbl})", str(r[idx])))
        if r[11] is None:
            flags.append("Jenis penarikan kosong")

        for kode_w, pesan, kol, nil in warn:
            res.issues.append(Issue(n, "warning", kode_w, pesan, kol, nil, kode))
        for i in err:
            i.branch_code = kode
        res.issues.extend(err)

        wajib = [r[2], r[3], r[4], r[5], r[6], r[8], r[9], r[10], r[11]]
        res.rows.append({
            "_ditolak": bool(err), "baris_no": n,
            "branch_code": kode if kode in branches else None,
            "tgl_input": tgl_input,
            "no_deposito": clean(r[4]), "no_deposito_norm": digits(r[4]),
            "nama_pemilik": clean(r[5]),
            "tgl_penempatan": tgl_tempat, "tgl_bilyet": as_date(r[7]), "tgl_pencairan": tgl_cair,
            "tenor_hari": tenor, "nominal": nominal,
            "jenis_pencairan": clean(r[10]), "jenis_penarikan": clean(r[11]),
            "data_tbo": clean(r[12]),
            "arus_dana": arus, "arus_keyakinan": yakin, "arus_manual": False,
            "nip_maker": clean(r[13]), "nip_checker": clean(r[14]), "nip_approver": clean(r[15]),
            "checker_eq_approver": bool(r[14] is not None and str(r[14]).strip() == str(r[15]).strip()),
            "catatan": catatan,
            "skor_lengkap": round(sum(1 for x in wajib if x is not None) / len(wajib), 3),
            "flags": flags,
        })

    wb.close()
    _tandai_duplikat(res, ("no_deposito_norm", "nominal", "tgl_pencairan"),
                     "Duplikat (nomor+nominal+tgl sama)", mode=mode_dup)
    return res


# --------------------------------------------------------------------------
# 3. PEMBUKAAN REKENING TBO - dari cabang
# --------------------------------------------------------------------------
SHEET_TBO = "Data TBO Pembukaan Rekening"
HEADER_ROW_TBO = 3

_KW_KEMBALI = re.compile(r"tempatkan kembali|penempatan kembali|jatuh tempo", re.I)
_PRODUK = [("deposito on call", "Deposito On Call"), ("oncall", "Deposito On Call"),
           ("doc", "Deposito On Call"), ("bundling", "Bundling"),
           ("giro", "Giro"), ("tabungan", "Tabungan"), ("deposito", "Deposito")]
_TIDAK_ADA = re.compile(r"^\s*(tidak\s*ada|tdk\s*ada|tidak\s*ad|-)\s*$", re.I)


def parse_tbo(path, branches: dict, settings: dict) -> ParseResult:
    res = ParseResult("tbo")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[SHEET_TBO] if SHEET_TBO in wb.sheetnames else wb.worksheets[0]

    for n, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if n <= HEADER_ROW_TBO or r is None or r[1] is None:
            continue
        res.raw.append((n, [str(x) if x is not None else None for x in r]))
        err = []

        kode = branch_code(r[2])
        if kode not in branches:
            err.append(Issue(n, "error", "cabang_tak_dikenal",
                             f"Kode cabang tidak dikenal: {r[2]}", "Kode Cabang", str(r[2])))
        tgl_input = as_date(r[3])
        if tgl_input is None:
            err.append(Issue(n, "error", "tgl_input_kosong", "Tanggal input kosong", "Tanggal (Input Data)"))

        # --- nominal: dua kolom IDR / valas sering tertukar ---
        flags = []
        idr = valas = None
        try:
            idr = as_number(r[9])
        except AmbiguousNumber as e:
            err.append(Issue(n, "error", "nominal_ambigu",
                             f"Nominal IDR tidak bisa dibaca tanpa menebak: {e}",
                             "Nominal Setoran Awal (IDR)", str(r[9])))
        try:
            valas = as_number(r[10])
        except AmbiguousNumber as e:
            res.issues.append(Issue(n, "warning", "valas_ambigu",
                                    f"Nominal mata uang lain tidak terbaca: {e}",
                                    "Nominal Setoran Awal (other currencies)", str(r[10])))

        if idr is not None and valas is not None and abs(idr - valas) < 0.005:
            # pola nyata: satu hari penuh baris mengisi kedua kolom dengan nilai sama
            flags.append("Kolom IDR dan valas diisi nilai sama")
            res.issues.append(Issue(n, "warning", "nominal_dobel",
                                    "Kolom IDR dan mata uang lain diisi nilai yang sama - "
                                    "kemungkinan salah kolom; dipakai sebagai IDR",
                                    "Nominal Setoran Awal", str(idr)))
            nominal, mata = idr, "IDR"
        elif idr is not None:
            nominal, mata = idr, "IDR"
        elif valas is not None:
            nominal, mata = valas, "NON-IDR"
            flags.append("Nominal hanya terisi di kolom mata uang lain")
        else:
            nominal, mata = None, "IDR"
            if not err:
                res.issues.append(Issue(n, "warning", "nominal_kosong",
                                        "Nominal setoran awal kosong", "Nominal Setoran Awal"))

        # --- CIF gabungan (joint account) ---
        cif_raw = clean(r[4])
        cif_gab = bool(cif_raw and "/" in cif_raw)

        # --- produk & tipe pembukaan dari kolom keterangan yang campur aduk ---
        ket = clean(r[17]) or ""
        produk = next((lbl for kw, lbl in _PRODUK if kw in ket.lower()), None)
        tgl_tempat = as_date(r[7])
        tipe = "Baru"
        if _KW_KEMBALI.search(ket) or (tgl_tempat and tgl_input and (tgl_input - tgl_tempat).days > 7):
            tipe = "Penempatan Kembali"
            flags.append("Terdeteksi penempatan kembali, bukan rekening baru")

        dok = clean(r[13])
        ada_tbo = bool(dok) and not _TIDAK_ADA.match(dok)

        if r[12] is None:
            flags.append("Jenis setoran kosong")

        for i in err + [x for x in res.issues if x.baris_no == n and x.branch_code is None]:
            i.branch_code = kode
        res.issues.extend(err)
        res.rows.append({
            "_ditolak": bool(err), "baris_no": n,
            "branch_code": kode if kode in branches else None,
            "tgl_input": tgl_input,
            "no_cif": cif_raw, "cif_gabungan": cif_gab,
            "no_rekening": clean(r[5]), "no_rekening_norm": digits(r[5]),
            "nama_pemilik": clean(r[6]),
            "tgl_penempatan": tgl_tempat, "tgl_jatuh_tempo": as_date(r[8]),
            "nominal": nominal, "mata_uang": mata,
            "jenis_rekening": clean(r[11]), "jenis_setoran": clean(r[12]),
            "jenis_produk": produk, "tipe_pembukaan": tipe,
            "dokumen_tbo": dok, "ada_tbo": ada_tbo,
            # status TBO selalu mulai Outstanding; dilengkapi lewat aplikasi
            "status_tbo": "Outstanding" if ada_tbo else "Dikecualikan",
            "nip_maker": clean(r[14]), "nip_checker": clean(r[15]), "nip_approver": clean(r[16]),
            "keterangan": ket or None,
            "flags": flags,
        })

    wb.close()
    return res


# --------------------------------------------------------------------------
# duplikat
# --------------------------------------------------------------------------
def _tandai_duplikat(res: ParseResult, kunci, label, mode="peringatan"):
    """
    Tandai baris dengan kunci identik. Baris pertama selalu dianggap sah.

    mode:
      'abaikan'    - tidak diperiksa sama sekali; semua baris masuk dan dihitung
      'peringatan' - baris kembar tetap masuk, ditandai, dikecualikan dari agregat
      'tolak'      - baris kembar ditolak, tidak masuk ke database
    """
    if mode == "abaikan":
        for row in res.rows:
            row.setdefault("is_duplikat", False)
            row.setdefault("dup_dikecualikan", False)
        return

    hard = (mode == "tolak")
    seen = {}
    for row in res.rows:
        k = tuple(row.get(x) for x in kunci)
        if any(v is None for v in k):
            row.setdefault("is_duplikat", False)
            row.setdefault("dup_dikecualikan", False)
            continue
        if k in seen:
            row["is_duplikat"] = True
            row["dup_dikecualikan"] = True
            row["flags"] = row.get("flags", []) + [label]
            seen[k]["is_duplikat"] = True
            if label not in seen[k]["flags"]:
                seen[k]["flags"] = seen[k]["flags"] + [label]
            res.issues.append(Issue(
                row["baris_no"], "error" if hard else "warning", "duplikat",
                f"{label}; sama dengan baris {seen[k]['baris_no']}", None, str(k[0]),
                row.get("branch_code")))
            if hard:
                row["_ditolak"] = True
        else:
            seen[k] = row
            row.setdefault("is_duplikat", False)
            row.setdefault("dup_dikecualikan", False)


# --------------------------------------------------------------------------
# master cabang
# --------------------------------------------------------------------------
def parse_master(path):
    """File 01. Kode dan Nama Cabang -> list dict siap upsert ke tabel branches."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    out = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r is None or len(r) < 3 or r[1] is None:
            continue
        kode = branch_code(r[1])
        nama = clean(r[2])
        if not kode or not nama:
            continue
        tipe = ("KC" if "(KC)" in nama else "KCP" if "(KCP)" in nama
                else "Pusat" if kode.startswith("000") else "Lainnya")
        out.append({"branch_code": kode, "branch_name": nama, "branch_type": tipe,
                    "region": _region(kode), "core_alias": None})
    wb.close()
    return out


_REGION = {"1": "Jakarta & sekitarnya", "2": "Jawa Barat", "3": "Sumatera",
           "4": "Jawa Timur, Bali & Indonesia Timur", "0": "Kantor Pusat"}


def _region(kode):
    return _REGION.get(kode[0], "Lainnya")


PARSERS = {"it_break": parse_it, "pencairan": parse_pencairan, "tbo": parse_tbo}

DESKRIPSI = {
    "it_break": "Data Transaksi dari IT (break deposito)",
    "pencairan": "Data Transaksi dari Cabang - Pencairan Deposito",
    "tbo": "Data Transaksi dari Cabang - Buka Rekening TBO",
}
