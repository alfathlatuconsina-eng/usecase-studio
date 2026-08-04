#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Import People Development data (Trainings + Evaluations) from an Excel file.

Usage:
    venv/bin/python import_people_excel.py <path/to/file.xlsx>

The workbook must have two sheets named exactly:
    "Trainings"   — one row per training program
    "Evaluations" — one row per evaluation result

Behaviour
---------
* Append-only: existing rows are never modified or deleted.
* A training is skipped if a row with the same name already exists (case-insensitive).
* An evaluation is skipped if one already exists for the same training_id.
* Evaluations reference trainings by the "Training Name" column; trainings that
  don't exist in the DB (and aren't being added in the same run) are skipped.

Run on VPS:
    cd /opt/pmo/backend
    source venv/bin/activate
    python import_people_excel.py /tmp/people_data.xlsx
"""

import sys
import datetime as dt
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not found — run: pip install openpyxl")

from sqlalchemy import select, func
from app import Session, PeopleTraining, PeopleEvaluation


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _str(v, default=""):
    if v is None:
        return default
    return str(v).strip()


def _int(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return None


def _float(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v))
    except (ValueError, TypeError):
        return None


def _date(v):
    """Accept YYYY-MM-DD strings or Excel date objects; return YYYY-MM-DD str."""
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s  # keep as-is if we can't parse it


def _headers(ws):
    """Return {lower-stripped-header: col_index} for row 1."""
    return {
        str(cell.value).strip().lower(): cell.column - 1
        for cell in ws[1]
        if cell.value is not None
    }


def _get(row, headers, *keys):
    """Try multiple header aliases; return first match or None."""
    for k in keys:
        idx = headers.get(k.lower())
        if idx is not None and idx < len(row):
            return row[idx]
    return None


# ---------------------------------------------------------------------------
# importers
# ---------------------------------------------------------------------------

def import_trainings(ws, session):
    headers = _headers(ws)
    existing = {
        n.lower()
        for n in session.scalars(select(PeopleTraining.name)).all()
    }

    added, skipped = 0, 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        name = _str(_get(row, headers, "name", "training name", "program name"))
        if not name:
            print(f"  [trainings] row {row_idx}: skipped — no name")
            skipped += 1
            continue

        if name.lower() in existing:
            print(f"  [trainings] row {row_idx}: skip (already exists) — {name}")
            skipped += 1
            continue

        session.add(PeopleTraining(
            name=name,
            category=_str(_get(row, headers, "category")),
            method=_str(_get(row, headers, "method", "delivery method")),
            organizer=_str(_get(row, headers, "organizer")),
            date_start=_date(_get(row, headers, "date_start", "start date", "date start")),
            date_end=_date(_get(row, headers, "date_end", "end date", "date end")),
            target_pax=_int(_get(row, headers, "target_pax", "target pax", "target")),
            actual_pax=_int(_get(row, headers, "actual_pax", "actual pax", "actual")),
            status=_str(_get(row, headers, "status"), default="Planned"),
            budget=_float(_get(row, headers, "budget")),
            realization=_float(_get(row, headers, "realization")),
            notes=_str(_get(row, headers, "notes")),
        ))
        existing.add(name.lower())
        print(f"  [trainings] row {row_idx}: added — {name}")
        added += 1

    return added, skipped


def import_evaluations(ws, session):
    headers = _headers(ws)

    # Build name→id map (includes rows just committed by import_trainings)
    name_to_id = {
        name.lower(): tid
        for name, tid in session.execute(
            select(PeopleTraining.name, PeopleTraining.id)
        ).all()
    }

    # Training IDs that already have an evaluation
    evaluated_ids = set(
        session.scalars(select(PeopleEvaluation.training_id)).all()
    )

    added, skipped = 0, 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        t_name = _str(_get(row, headers, "training name", "training", "program name"))
        if not t_name:
            print(f"  [evaluations] row {row_idx}: skipped — no training name")
            skipped += 1
            continue

        tid = name_to_id.get(t_name.lower())
        if tid is None:
            print(f"  [evaluations] row {row_idx}: skipped — training not found: '{t_name}'")
            skipped += 1
            continue

        if tid in evaluated_ids:
            print(f"  [evaluations] row {row_idx}: skip (evaluation already exists) — {t_name}")
            skipped += 1
            continue

        session.add(PeopleEvaluation(
            training_id=tid,
            reaction_score=_float(_get(row, headers,
                "reaction_score", "reaction score", "reaction", "l1 score")),
            learning_score=_float(_get(row, headers,
                "learning_score", "learning score", "learning", "l2 score")),
            respondents=_int(_get(row, headers,
                "respondents", "respondent", "num respondents", "jumlah")),
            notes=_str(_get(row, headers, "notes")),
        ))
        evaluated_ids.add(tid)
        print(f"  [evaluations] row {row_idx}: added — {t_name}")
        added += 1

    return added, skipped


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python import_people_excel.py <path/to/file.xlsx>")

    path = Path(sys.argv[1])
    if not path.exists():
        sys.exit(f"File not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = [s.lower() for s in wb.sheetnames]

    total_added = 0

    with Session() as session:
        # --- Trainings ---
        tr_sheet = next(
            (wb[s] for s in wb.sheetnames if s.lower() == "trainings"), None
        )
        if tr_sheet:
            print("\n=== Importing Trainings ===")
            a, s = import_trainings(tr_sheet, session)
            print(f"  Result: {a} added, {s} skipped")
            total_added += a
        else:
            print("  No 'Trainings' sheet found — skipping")

        # --- Evaluations ---
        ev_sheet = next(
            (wb[s] for s in wb.sheetnames if s.lower() == "evaluations"), None
        )
        if ev_sheet:
            print("\n=== Importing Evaluations ===")
            a, s = import_evaluations(ev_sheet, session)
            print(f"  Result: {a} added, {s} skipped")
            total_added += a
        else:
            print("  No 'Evaluations' sheet found — skipping")

        session.commit()

    print(f"\nDone. Total rows inserted: {total_added}")


if __name__ == "__main__":
    main()
