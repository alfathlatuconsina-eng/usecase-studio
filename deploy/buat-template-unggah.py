#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Membuat TIGA berkas template unggahan Branch Ops, berisi contoh data.

    py -3 deploy\\buat-template-unggah.py

Hasilnya di folder contoh/:
    Template-01-Break-Deposito.xlsx
    Template-02-Pencairan-Deposito.xlsx
    Template-03-Data-TBO.xlsx

KENAPA berkas ini ada, bukan sekadar template yang dibuat sekali lalu
disimpan: ketiga parser membaca sel berdasarkan POSISI (r[0], r[1], ...),
bukan berdasarkan judul kolom. Jadi template harus selalu sama persis
dengan pemetaan di ingest.py. Dengan skrip, template bisa dibuat ulang
setiap kali parser berubah - dan skrip ini SEKALIGUS mengujinya:
di akhir, tiap berkas dibaca balik oleh parser aslinya, dan jumlah baris
diterima/ditolak dicetak. Kalau ada yang ditolak, templatenya salah.

Nama nasabah di contoh sengaja fiktif.
"""
import datetime
import pathlib
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

AKAR = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(AKAR / "backend"))

KELUAR = AKAR / "contoh"
KELUAR.mkdir(exist_ok=True)

JUDUL_FILL = PatternFill("solid", fgColor="D9E1F2")
BARU_FILL = PatternFill("solid", fgColor="FFF2CC")   # kolom tambahan Agu 2026
ABAI_FILL = PatternFill("solid", fgColor="F2F2F2")   # kolom tidak dibaca parser
TEBAL = Font(bold=True)

# Kode cabang HARUS sudah ada di branchops_branches, kalau tidak seluruh
# baris ditolak dengan "Kode cabang tidak dikenal".
CABANG = ["01006", "01001", "01008", "01003", "01004"]


def tulis(ws, baris_judul, judul, catatan, baris_data, lebar_khusus=()):
    """Judul kolom + baris data. Data mulai TEPAT di baris berikutnya.

    Keterangan tipe TIDAK ditulis sebagai baris di lembar data. Percobaan
    pertama menaruhnya di bawah judul, dan parser membacanya sebagai baris
    data — tiap template menghasilkan satu baris ditolak berisi teks
    'wajib, tanggal' dan sejenisnya. Parser tidak tahu mana baris hiasan;
    ia membaca semua yang ada setelah HEADER_ROW.

    Keterangannya pindah ke lembar PETUNJUK, lewat kolom_tabel()."""
    for i, j in enumerate(judul, start=1):
        c = ws.cell(row=baris_judul, column=i, value=j)
        c.font = TEBAL
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.fill = (BARU_FILL if j.startswith("* ") else
                  ABAI_FILL if j.startswith("(") else JUDUL_FILL)
    for b, baris in enumerate(baris_data):
        for i, v in enumerate(baris, start=1):
            ws.cell(row=baris_judul + 1 + b, column=i, value=v)
    for i in range(1, len(judul) + 1):
        ws.column_dimensions[get_column_letter(i)].width = lebar_khusus[i - 1] \
            if i - 1 < len(lebar_khusus) else 17
    ws.freeze_panes = ws.cell(row=baris_judul + 1, column=1)


def kolom_tabel(judul, catatan):
    """Daftar kolom untuk lembar PETUNJUK: posisi, indeks parser, tipe."""
    baris = ["DAFTAR KOLOM",
             "",
             "Kolom Excel | indeks parser | judul | keterangan",
             "-" * 96]
    for i, (j, k) in enumerate(zip(judul, catatan)):
        baris.append(f"  {get_column_letter(i + 1):>2}  |  r[{i:2d}]  |  "
                     f"{j:<32} | {k}")
    baris += ["", "Kolom bertanda * adalah tambahan Agustus 2026.",
              "Kolom berjudul (tidak dibaca) memang dilewati parser.", ""]
    return baris


def petunjuk(wb, teks):
    ws = wb.create_sheet("PETUNJUK")
    ws.column_dimensions["A"].width = 110
    for i, t in enumerate(teks, start=1):
        c = ws.cell(row=i, column=1, value=t)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if t and not t.startswith(" ") and t == t.upper() and len(t) > 3:
            c.font = TEBAL
    return ws


UMUM = [
    "ATURAN YANG BERLAKU UNTUK KETIGA TEMPLATE",
    "",
    "1. Parser membaca sel berdasarkan POSISI kolom, bukan judulnya.",
    "   Judul boleh diganti bahasanya; URUTAN KOLOM TIDAK BOLEH BERUBAH.",
    "   Menyisipkan kolom di tengah menggeser semua kolom sesudahnya, dan",
    "   data akan masuk ke kolom yang salah TANPA pesan galat.",
    "",
    "2. Kolom baru selalu ditambahkan di PALING KANAN, tidak pernah di tengah.",
    "",
    "3. Baris judul tidak boleh dipindah. Data dimulai pada baris yang sudah",
    "   ditentukan di tiap template (lihat catatan di lembar ini).",
    "",
    "4. Kode cabang harus sudah terdaftar di Master Cabang. Kalau belum,",
    "   seluruh barisnya ditolak dengan 'Kode cabang tidak dikenal'.",
    "   Unggah '01. Kode dan Nama Cabang.xlsx' lebih dulu di Langkah 0.",
    "",
    "5. Sel tanggal harus benar-benar bertipe TANGGAL di Excel, bukan teks.",
    "   Parser tidak menebak-nebak format; teks '01/08/2026' dibaca kosong.",
    "",
    "6. Kolom berlatar KUNING adalah tambahan Agustus 2026. Boleh dikosongkan;",
    "   berkas lama tanpa kolom itu tetap bisa diunggah.",
    "   Kolom berlatar ABU-ABU tidak dibaca parser - biarkan apa adanya.",
    "",
    "7. Unggah menghasilkan batch berstatus 'draft'. Data baru muncul di",
    "   dashboard setelah ditekan 'Komit' di tab Unggah.",
    "",
]


# ==========================================================================
# 1. BREAK DEPOSITO  (parse_it)  - sheet "Sheet1", judul baris 1, data baris 2
# ==========================================================================
def break_deposito():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    judul = [
        "Kode Cabang", "Nama Cabang (core)", "Saldo", "(tidak dibaca)",
        "Tanggal Penempatan", "Tanggal Jatuh Tempo", "Tanggal Break",
        "Waktu Awal", "Waktu Akhir", "Rekening Pendebetan", "Nama Pemilik",
        "Nominal", "Penalti", "Mata Uang", "Rate", "Rekening Pencairan",
        "Nama Pencairan", "(tidak dibaca)", "CS ID", "CS Nama",
        "(tidak dibaca)", "FLM1 Nama", "Teller ID", "Teller Nama",
        "FLM2 ID", "FLM2 Nama",
    ]
    catatan = [
        "wajib, 4-5 digit", "teks", "angka", "", "tanggal", "tanggal",
        "wajib, tanggal", "jam hh:mm:ss", "jam hh:mm:ss", "teks",
        "teks, disamarkan", "wajib, angka", "angka", "IDR / lainnya",
        "angka desimal", "teks", "teks, disamarkan", "", "teks", "teks",
        "", "teks", "teks", "teks", "teks", "teks",
    ]

    d = datetime.date
    t = datetime.time
    data = []
    contoh = [
        ("01006", "CAB.JAKARTA-GREEN GARDEN", 3_450_000_000, d(2024, 8, 16),
         d(2026, 8, 16), d(2026, 8, 3), t(9, 36, 8), t(9, 36, 30),
         "300010002826028", "BUDI HARTONO", 338_944_580.82, 267_178.08, 0.06),
        ("01001", "CAB.JAKARTA-WISMA BUMIPUTERA", 1_200_000_000, d(2025, 2, 1),
         d(2026, 8, 20), d(2026, 8, 3), t(10, 15, 0), t(10, 15, 42),
         "300010002826029", "PT CONTOH SEJAHTERA", 1_200_000_000, 900_000, 0.055),
        ("01008", "CAB.JAKARTA-PURI INDAH", 500_000_000, d(2026, 1, 12),
         d(2026, 9, 12), d(2026, 8, 4), t(14, 2, 11), t(14, 3, 0),
         "300010002826030", "SITI RAHAYU", 500_000_000, 250_000, 0.0575),
        ("01003", "CAB.JAKARTA-WOLTER", 750_000_000, d(2025, 11, 3),
         d(2026, 11, 3), d(2026, 8, 4), t(16, 40, 5), t(16, 41, 20),
         "300010002826031", "DEWI ANGGRAINI", 748_500_000, 601_500, 0.0625),
        ("01004", "CAB.JAKARTA-ROXY", 250_000_000, d(2026, 3, 20),
         d(2026, 8, 5), d(2026, 8, 5), t(8, 5, 44), t(8, 6, 10),
         "300010002826032", "CV MITRA CONTOH", 250_000_000, 0, 0.05),
    ]
    for i, (kode, nama_cab, saldo, tp, jt, br, wa, wb_, rek, nama,
            nom, pen, rate) in enumerate(contoh):
        data.append([
            kode, nama_cab, saldo, 0, tp, jt, br, wa, wb_, rek, nama,
            nom, pen, "IDR", rate, "209010000331601", nama, "Y",
            f"MNC24134{i:02d}", "PETUGAS CS CONTOH", f"MNC04101{i:02d}",
            "PENYELIA CONTOH", f"41010{i}", "TELLER CONTOH",
            f"41014{i}", "PENYELIA 2 CONTOH",
        ])

    tulis(ws, 1, judul, catatan, data, lebar_khusus=[13, 30] + [17] * 24)
    petunjuk(wb, UMUM + kolom_tabel(judul, catatan) + [
        "KHUSUS BREAK DEPOSITO",
        "",
        "Nama lembar : Sheet1  (kalau berbeda, lembar PERTAMA yang dipakai)",
        "Baris judul : 1",
        "Data mulai  : baris 2",
        "Baris dilewati bila kolom A (Kode Cabang) kosong.",
        "",
        "Kolom wajib : Kode Cabang, Tanggal Break, Nominal.",
        "Kolom 4, 18 dan 21 tidak dibaca parser - di ekspor asli isinya",
        "duplikat dari kolom lain.",
        "",
        "Kolom bebas berikutnya kalau perlu menambah field: kolom ke-27 (r[26]).",
        "",
        "Tanggal Break di luar rentang wajar TIDAK ditolak, hanya diberi",
        "peringatan. Batch 27 (Agustus 2026) lolos dengan seluruh tanggal",
        "1984-05-24 karena itu - periksa hasil unggahan sebelum Komit.",
    ])
    p = KELUAR / "Template-01-Break-Deposito.xlsx"
    wb.save(p)
    return p


# ==========================================================================
# 2. PENCAIRAN DEPOSITO  (parse_pencairan)
#    sheet "Pencairan Deposito - olah", judul baris 3, data baris 4
# ==========================================================================
def pencairan():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pencairan Deposito - olah"
    ws["A1"] = "PENCAIRAN DEPOSITO — laporan cabang"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Baris 1-3 dilewati parser. Judul di baris 3, data mulai baris 4."

    judul = [
        "(tidak dibaca)", "No.", "Kode & Nama Cabang", "Tanggal (Input Data)",
        "No. Deposito", "Nama Pemilik", "Tanggal Penempatan", "Tanggal Bilyet",
        "Tanggal Pencairan", "Nominal", "Jenis Pencairan", "Jenis Penarikan",
        "Data TBO", "NIP Maker", "NIP Checker", "NIP Approver", "Catatan",
        "* No. CIF", "* No. Rekening", "* Target Pemenuhan TBO",
    ]
    catatan = [
        "", "wajib diisi, nomor urut", "wajib, mis. 1006 - GREEN GARDEN (KCP)",
        "wajib, tanggal", "teks", "teks, disamarkan", "tanggal", "tanggal",
        "wajib, tanggal", "wajib, angka", "teks", "teks",
        "teks — isi = baris dilacak sebagai TBO", "9 digit", "9 digit",
        "9 digit", "teks bebas", "baru, boleh kosong", "baru, boleh kosong",
        "baru, TANGGAL, boleh kosong",
    ]

    d = datetime.date
    data = [
        [None, 1, "1006 - GREEN GARDEN (KCP)", d(2026, 8, 3), "DEP-2026-0001",
         "BUDI HARTONO", d(2026, 2, 3), d(2026, 2, 3), d(2026, 8, 3),
         500_000_000, "Sesuai Jatuh Tempo", "Pemindahbukuan",
         "Form Penempatan belum lengkap", "202500304", "200923685", "200923686",
         "Nasabah hadir di cabang", "959001", "0100612345671", d(2026, 8, 20)],
        [None, 2, "1001 - JKT WISMA BUMIPUTERA (KCP)", d(2026, 8, 3), "DEP-2026-0002",
         "PT CONTOH SEJAHTERA", d(2026, 7, 4), None, d(2026, 8, 3),
         3_000_000_000, "Dipercepat dari Jatuh Tempo", "Transfer",
         None, "202500305", "200923685", "200923686",
         "Ditempatkan kembali dengan nominal yg sama", None, None, None],
        [None, 3, "1008 - PURI INDAH (KCP)", d(2026, 8, 4), "DEP-2026-0003",
         "SITI RAHAYU", d(2026, 8, 3), d(2026, 8, 3), d(2026, 8, 4),
         250_000_000, "Sesuai Jatuh Tempo", "Tunai",
         "KTP belum diperbarui", "202500306", "200923687", "200923688",
         "Rollover otomatis", "959002", "0100812345672", d(2026, 8, 15)],
        [None, 4, "1003 - WOLTER (KCP)", d(2026, 8, 4), "DEP-2026-0004",
         "DEWI ANGGRAINI", d(2026, 5, 4), d(2026, 5, 4), d(2026, 8, 4),
         750_000_000, "Sesuai Jatuh Tempo", "Pemindahbukuan",
         "tidak ada", "202500307", "200923689", "200923690",
         None, "959003", "0100312345673", None],
        [None, 5, "1004 - ROXY (KCP)", d(2026, 8, 5), "DEP-2026-0005",
         "CV MITRA CONTOH", d(2026, 8, 4), None, d(2026, 8, 5),
         1_000_000_000, "Dipercepat dari Jatuh Tempo", "Transfer",
         "Spesimen tanda tangan", "202500308", "200923691", "200923691",
         "Checker dan approver sama", "959004", "0100412345674", d(2026, 7, 30)],
    ]

    tulis(ws, 3, judul, catatan, data,
          lebar_khusus=[13, 6, 32, 16, 17, 26, 16, 15, 17, 16, 22, 17, 28,
                        12, 12, 12, 30, 13, 17, 20])
    petunjuk(wb, UMUM + kolom_tabel(judul, catatan) + [
        "KHUSUS PENCAIRAN DEPOSITO",
        "",
        "Nama lembar : Pencairan Deposito - olah",
        "Baris judul : 3",
        "Data mulai  : baris 5 pada template ini (baris 4 dipakai keterangan",
        "              tipe; boleh dihapus, data lalu mulai baris 4).",
        "Baris dilewati bila kolom B (No.) kosong.",
        "",
        "Kolom wajib : Kode & Nama Cabang, Tanggal (Input Data),",
        "              Tanggal Pencairan, Nominal.",
        "",
        "TIGA KOLOM BARU (Agustus 2026), semuanya di paling kanan:",
        "  kolom 18  No. CIF                 - boleh kosong",
        "  kolom 19  No. Rekening            - boleh kosong",
        "  kolom 20  Target Pemenuhan TBO    - boleh kosong, harus TANGGAL",
        "",
        "Kolom bebas berikutnya: kolom ke-21 (r[20]).",
        "",
        "KOLOM 'Data TBO' MENENTUKAN BANYAK HAL:",
        "  - terisi  -> baris dilacak sebagai TBO, status awal Outstanding,",
        "               muncul di Beranda, dan bisa disunting lewat tombol Ubah.",
        "  - kosong, atau berisi 'tidak ada' / 'tdk ada' / '-'",
        "               -> status Dikecualikan, tidak dilacak.",
        "Mengisi Target Pemenuhan TBO tanpa mengisi Data TBO menghasilkan",
        "peringatan: barisnya tidak akan pernah dihitung terlambat.",
        "",
        "Arus dana (Arus Keluar / Rollover / Penempatan Kembali) DIHITUNG",
        "sendiri dari tenor dan kata kunci di kolom Catatan - tidak ada",
        "kolomnya di berkas ini.",
    ])
    p = KELUAR / "Template-02-Pencairan-Deposito.xlsx"
    wb.save(p)
    return p


# ==========================================================================
# 3. DATA TBO  (parse_tbo)
#    sheet "Data TBO Pembukaan Rekening", judul baris 3, data baris 4
# ==========================================================================
def tbo():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data TBO Pembukaan Rekening"
    ws["A1"] = "DATA TBO — pembukaan rekening dengan dokumen TBO"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Baris 1-3 dilewati parser. Judul di baris 3, data mulai baris 4."

    judul = [
        "(tidak dibaca)", "No.", "Kode & Nama Cabang", "Tanggal (Input Data)",
        "No. CIF", "No. Rekening", "Nama Pemilik", "Tanggal Penempatan",
        "Tanggal Jatuh Tempo", "Nominal Setoran Awal (IDR)",
        "Nominal Setoran Awal (valas)", "Jenis Rekening", "Jenis Setoran",
        "Dokumen TBO", "NIP Maker", "NIP Checker", "NIP Approver",
        "Keterangan", "* Target Pemenuhan TBO",
    ]
    catatan = [
        "", "wajib diisi, nomor urut", "wajib, mis. 1100 - SOEPOMO (KC)",
        "wajib, tanggal", "teks; '/' = CIF gabungan", "teks",
        "teks, disamarkan", "tanggal", "tanggal", "angka",
        "angka, isi hanya bila non-IDR", "Perorangan / Perusahaan",
        "Tunai / Transfer / ...", "teks — kosong atau 'tidak ada' = tanpa TBO",
        "9 digit", "9 digit", "9 digit",
        "teks; kata 'deposito'/'on call' menentukan jenis produk",
        "baru, TANGGAL, boleh kosong",
    ]

    d = datetime.date
    data = [
        [None, 1, "1006 - GREEN GARDEN (KCP)", d(2026, 8, 3), "959933",
         "300010003244775", "PT INOVASI CONTOH", d(2026, 8, 3), d(2026, 9, 3),
         1_200_000_000, None, "Perusahaan (Non Perorangan)", "Transfer",
         "Form Penempatan", "202404801", "201127091", "201127092",
         "Deposito", d(2026, 8, 17)],
        [None, 2, "1001 - JKT WISMA BUMIPUTERA (KCP)", d(2026, 8, 3), "959934",
         "300010003244776", "BUDI HARTONO", d(2026, 8, 3), d(2026, 11, 3),
         500_000_000, None, "Perorangan", "Tunai",
         "KTP dan NPWP", "202404802", "201127093", "201127094",
         "Deposito baru", d(2026, 8, 10)],
        [None, 3, "1008 - PURI INDAH (KCP)", d(2026, 8, 4), "959935/959936",
         "300010003244777", "SITI RAHAYU", d(2026, 8, 4), d(2026, 8, 11),
         250_000_000, None, "Perorangan", "Pemindahbukuan",
         "Form OR", "202404803", "201127095", "201127096",
         "Deposito On Call", d(2026, 8, 8)],
        [None, 4, "1003 - WOLTER (KCP)", d(2026, 8, 4), "959937",
         "300010003244778", "DEWI ANGGRAINI", d(2026, 8, 4), d(2027, 8, 4),
         750_000_000, None, "Perorangan", "Transfer",
         "tidak ada", "202404804", "201127097", "201127098",
         "Tempatkan kembali dari jatuh tempo", None],
        [None, 5, "1004 - ROXY (KCP)", d(2026, 8, 5), "959938",
         "300010003244779", "CV MITRA CONTOH", d(2026, 8, 5), d(2026, 9, 5),
         None, 25_000, "Perusahaan (Non Perorangan)", "Transfer",
         "Form Penempatan dan Spesimen", "202404805", "201127099", "201127100",
         "Deposito valas", d(2026, 8, 19)],
    ]

    tulis(ws, 3, judul, catatan, data,
          lebar_khusus=[13, 6, 32, 16, 16, 19, 26, 16, 17, 20, 20, 24, 17,
                        28, 12, 12, 12, 30, 20])
    petunjuk(wb, UMUM + kolom_tabel(judul, catatan) + [
        "KHUSUS DATA TBO",
        "",
        "Nama lembar : Data TBO Pembukaan Rekening",
        "Baris judul : 3",
        "Data mulai  : baris 5 pada template ini (baris 4 dipakai keterangan",
        "              tipe; boleh dihapus, data lalu mulai baris 4).",
        "Baris dilewati bila kolom B (No.) kosong.",
        "",
        "Kolom wajib : Kode & Nama Cabang, Tanggal (Input Data).",
        "",
        "SATU KOLOM BARU (Agustus 2026), di paling kanan:",
        "  kolom 19  Target Pemenuhan TBO - boleh kosong, harus TANGGAL",
        "",
        "Kolom bebas berikutnya: kolom ke-20 (r[19]).",
        "",
        "KOLOM 'Dokumen TBO' MENENTUKAN STATUS AWAL:",
        "  - terisi  -> ada_tbo = benar, status awal Outstanding.",
        "  - kosong, atau 'tidak ada' / 'tdk ada' / '-'",
        "               -> status Dikecualikan, tidak dilacak.",
        "",
        "Nominal: isi kolom IDR ATAU kolom valas, jangan keduanya. Mengisi",
        "keduanya dengan nilai sama memicu peringatan 'nominal_dobel'.",
        "",
        "Status TBO dan Tanggal TBO Lengkap TIDAK ada di berkas ini - itu",
        "diisi lewat aplikasi (tombol Ubah / Tandai lengkap), karena yang",
        "tahu dokumen sudah lengkap adalah kantor pusat, bukan cabang.",
    ])
    p = KELUAR / "Template-03-Data-TBO.xlsx"
    wb.save(p)
    return p


# ==========================================================================
# Uji balik: baca template dengan parser aslinya
# ==========================================================================
def _muat_ingest():
    """Muat ingest.py LANGSUNG, tanpa lewat paket branchops.

    branchops/__init__.py mengimpor Flask; skrip ini tidak perlu Flask dan
    harus tetap bisa dijalankan di lingkungan yang belum memasangnya."""
    import importlib.util
    import types
    pkg = types.ModuleType("branchops")
    pkg.__path__ = [str(AKAR / "backend" / "branchops")]
    sys.modules.setdefault("branchops", pkg)
    spec = importlib.util.spec_from_file_location(
        "branchops.ingest", AKAR / "backend" / "branchops" / "ingest.py")
    m = importlib.util.module_from_spec(spec)
    sys.modules["branchops.ingest"] = m
    spec.loader.exec_module(m)
    return m


def uji(berkas):
    ingest = _muat_ingest()

    branches = {k: {"branch_code": k} for k in CABANG}
    settings = {"jam_operasional_mulai": 8, "jam_operasional_selesai": 16,
                "rollover_tenor_hari": 1, "validasi_nip": 0,
                "validasi_duplikat": "abaikan"}
    fn = {"it_break": ingest.parse_it, "pencairan": ingest.parse_pencairan,
          "tbo": ingest.parse_tbo}

    print("\n" + "=" * 68)
    print("UJI BALIK — template dibaca oleh parser yang sesungguhnya")
    print("=" * 68)
    semua_lolos = True
    for jenis, path in berkas.items():
        res = fn[jenis](str(path), branches, settings)
        ditolak = [r for r in res.rows if r["_ditolak"]]
        err = [i for i in res.issues if i.severity == "error"]
        warn = [i for i in res.issues if i.severity == "warning"]
        ok = not ditolak and not err
        semua_lolos &= ok
        print(f"\n  {path.name}")
        print(f"    jenis          : {jenis}")
        print(f"    baris terbaca  : {len(res.rows)}")
        print(f"    baris ditolak  : {len(ditolak)}   {'OK' if not ditolak else '<-- PERIKSA'}")
        print(f"    error          : {len(err)}")
        print(f"    warning        : {len(warn)}")
        for w in warn[:4]:
            print(f"       - baris {w.baris_no}: {w.pesan[:74]}")
        if res.rows:
            r0 = res.rows[0]
            baru = {"it_break": [],
                    "pencairan": ["no_cif", "no_rekening",
                                  "target_pemenuhan_tbo", "status_tbo"],
                    "tbo": ["target_pemenuhan_tbo", "status_tbo"]}[jenis]
            if baru:
                print("    kolom baru terbaca pada baris pertama:")
                for k in baru:
                    print(f"       {k:22s} = {r0.get(k)!r}")
    print("\n" + ("SEMUA TEMPLATE LOLOS" if semua_lolos
                  else "ADA TEMPLATE YANG DITOLAK — perbaiki sebelum dipakai"))
    return semua_lolos


if __name__ == "__main__":
    berkas = {"it_break": break_deposito(),
              "pencairan": pencairan(),
              "tbo": tbo()}
    for j, p in berkas.items():
        print(f"dibuat: {p}")
    sys.exit(0 if uji(berkas) else 1)
